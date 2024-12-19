import logging

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)


class Config:
    def __init__(self) -> None:
        from pathlib import Path

        self.input_name = "Profile 2025 GD.xlsx"
        self.input_path = Path.cwd() / self.input_name


def df_transform(values):
    import pandas as pd
    from datetime import datetime, time

    df = pd.DataFrame(values)

    headers = df.iloc[0].tolist()
    headers = [
        item if isinstance(item, time) else item.time()
        for item in headers
        if isinstance(item, time) or isinstance(item, datetime)
    ]
    logging.debug(f"headers:\n{headers}")

    dates = df.iloc[1:, 0].tolist()
    dates = [item.date() for item in dates if isinstance(item, datetime)]
    logging.debug(f"dates:\n{dates}")

    datetime_list = []
    value_list = []

    for row_index in range(len(dates)):
        row = df.iloc[row_index + 1]

        for col_index in range(len(headers)):
            value = row.iloc[col_index + 1]

            datetime_list.append(datetime.combine(dates[row_index], headers[col_index]))
            value_list.append(value)

    new_df = pd.DataFrame({"datetime": datetime_list, "value": value_list})
    new_df["datetime"] = pd.to_datetime(new_df["datetime"])
    new_df = new_df.sort_values(by="datetime")

    return new_df


def save_df_to_sheet(workbook, sheet_name, df):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo

    new_sheet = workbook.create_sheet(sheet_name)

    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=1
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)

            col_letter = new_sheet.cell(row=r_idx, column=c_idx).column_letter
            new_sheet.column_dimensions[col_letter].width = max(
                new_sheet.column_dimensions[col_letter].width, len(str(value)) + 2
            )

    rows, cols = df.shape
    table_range = f"A1:{chr(64 + cols)}{rows + 1}"

    table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_range)

    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style

    new_sheet.add_table(table)


def main() -> None:
    from openpyxl import load_workbook

    logging.info("Start")
    config = Config()
    workbook = load_workbook(config.input_path)
    sheets_names = workbook.sheetnames

    for sheet_name in sheets_names:
        sheet = workbook[sheet_name]
        df = df_transform(sheet.values)
        save_df_to_sheet(workbook, f"{sheet_name} ISO", df)

    workbook.save(config.input_path)
    logging.info("End")


if __name__ == "__main__":
    main()
